import os
from openpyxl import load_workbook

# Define las rutas de los ficheros
libro1_path = "C:\\Users\\lhida\\OneDrive\\Desktop\\New folder\\Nucleus---eWallet\\Libro1.xlsx"
libro2_path = "C:\\Users\\lhida\\OneDrive\\Desktop\\New folder\\Nucleus---eWallet\\Libro2.xlsx"

# Abre los libros de Excel existentes
libro1 = load_workbook(libro1_path)
libro2 = load_workbook(libro2_path)

# Selecciona la hoja activa de cada libro
hoja1 = libro1.active  
hoja2 = libro2.active  


datos_libro2 = {}
for fila in hoja2.iter_rows(min_row=2, values_only=True):  # Se salta la cabecera
    id_valor = fila[0]  # ID en la columna A
    valor = fila[1]    # Valor en la columna B
    valor_desperdicio = fila[2]
    if id_valor is not None:  # Asegúrate de que el ID no esté vacío
        datos_libro2[id_valor] = valor

# Función para actualizar el estatus
def FuncionStatus(estatus_libro1, fila):
    if estatus_libro1 == "EX":
        fila[2].value = "IM"
    elif estatus_libro1 == "IM":
        fila[2].value = "LA"
    elif estatus_libro1 == "LA":
        fila[2].value = "CO"
    elif estatus_libro1 == "CO":
        fila[2].value = "SL"
    elif estatus_libro1 is None or estatus_libro1 == "":  # Si está vacío
        fila[2].value = "EX"


# def Desperdicio(valor_desperdicio, desperdicio_libro1):
#     if valor_desperdicio == "" :
#         desperdicio_libro1 += valor_desperdicio

#     else:
#         valor_desperdicio = "no hay desperdicio"


# Actualiza Libro1
for fila in hoja1.iter_rows(min_row=2):  # Se salta la cabecera
    id_libro1 = fila[0].value  # ID en la columna A
    valor_libro1 = fila[1].value  # Valor en la columna B
    estatus_libro1 = fila[2].value  # Estatus en la columna C
    desperdicio_libro1 = fila[3].value


    if id_libro1 in datos_libro2:  # Compara el ID con los de Libro2
        nuevo_valor = datos_libro2[id_libro1]
        
        # Actualiza el valor si es necesario
        if valor_libro1 is None:  # Si el valor en Libro1 está vacío
            fila[1].value = nuevo_valor  # Asigna el valor de Libro2
        else:
            # Asegúrate de que valor_libro1 es un número antes de sumar
            if isinstance(valor_libro1, (int, float)) and isinstance(nuevo_valor, (int, float)):
                fila[1].value += nuevo_valor  # Suma el valor de Libro2 al de Libro1
            else:
                print(f"Valor no numérico en fila {fila[0].row}, columna B: {valor_libro1}. No se sumará.")
                
        FuncionStatus(estatus_libro1, fila)

        # Desperdicio(desperdicio_libro1,valor_desperdicio)


libro1.save(libro1_path)


libro1.close()
libro2.close()

# Abre el archivo actualizado utilizando la aplicación predeterminada del sistema
os.startfile(libro1_path)
